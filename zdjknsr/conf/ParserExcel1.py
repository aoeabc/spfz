import openpyxl
from openpyxl.styles import Border,Side,Font

class ParserExcel:

    def __init__(self,excelfile):
        self.work_book = openpyxl.load_workbook(excelfile)
        self.excelFile = None
        self.font = Font(color=None)
        self.RGBDict = {'red': 'FFFF3030', 'green':'FF008B00'}

    # def loadWorkBook(self,excelfile):
    #     # 根据路径获取表格对象
    #     try:
    #         self.work_book =  openpyxl.load_workbook(excelfile)
    #     except Exception as e:
    #         raise e
    #     self.excelFile = excelfile
    #     return self.work_book

    def getSheetByName(self, sheet_name):
        #  根据sheet表名获取工作表对象
        try:
            sheet = self.work_book[sheet_name]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheet_index):
        #  根据sheet表索引获取工作表对象
        try:
            sheet_name = self.work_book.sheetnames[sheet_index]
        except Exception as e:
            raise e
        sheet = self.work_book[sheet_name]
        return sheet

    def getRowsNumber(self, sheet):
        #  获取工作表内容行数
        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取工作表内容列数
        return sheet.max_column

    def getStartRowNunber(self, sheet):
        #  获取表格中的开始行号
        return sheet.min_row

    def getStartColNunber(self, sheet):
        #  获取表格中的开始列号
        return sheet.min_column

    def getRow(self, sheet, rownumber):
        # 获取某一行的内容
        return list(sheet.rows)[rownumber]

    def getCol(self, sheet, colnumber):
        # 获取某一行的内容
        return list(sheet.columns)[colnumber]
    
    def getCellOfValue(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 根据索引获取单元格的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate).value
            except Exception as e:
                raise e
        if coordinate == None and rowNo != None and colsNo != None:
            try:
                return sheet.cell(row=rowNo, colnum=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("单元格找不到")

    def getCellOfObject(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 根据索引获取单元格的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate = coordinate)
            except Exception as e:
                raise e
        if coordinate == None and rowNo != None and colsNo != None:
            try:
                return sheet.cell(row=rowNo, colnum=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("单元格找不到")

    def writeCell(self, sheet, contant, coordinate = None, rowNo = None, colsNo = None, style = None):
        if coordinate != None:
            try:
                sheet.cell(coordinate = coordinate).value = contant
                if style != None:
                    sheet.cell(coordinate = coordinate).font = Font(color=self.RGBDict[style])
                self.work_book.save(self.excelFile)
            except Exception as e:
                raise e

        elif coordinate == None and rowNo != None and colsNo != None:
            try:
                sheet.cell(row = rowNo, column = colsNo).value = contant
                if style != None:
                    sheet.cell(row = rowNo, column = colsNo).font = Font(color=self.RGBDict[style])
                self.work_book.save(self.excelFile)
            except Exception as e:
                raise e

        else:
            raise Exception("单元格找不到")